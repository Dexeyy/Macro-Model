from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from pathlib import Path
from typing import List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import Chart


@dataclass
class ChartInfo:
    sheet: str
    title: str
    series_titles: List[str]
    series_formulas: List[str]


def _safe_title(ch: Chart) -> str:
    try:
        if ch.title and hasattr(ch.title, "tx") and hasattr(ch.title.tx, "rich"):
            return str(ch.title.tx.rich.p[0].r[0].t)
        if ch.title and hasattr(ch.title, "tx") and hasattr(ch.title.tx, "strRef"):
            return str(ch.title.tx.strRef.f)
        if ch.title:
            return str(ch.title)
    except Exception:
        pass
    return ""


def _series_meta(ch: Chart) -> Tuple[List[str], List[str]]:
    titles: List[str] = []
    formulas: List[str] = []
    try:
        for s in getattr(ch, "series", []) or []:
            # Title
            try:
                if getattr(s, "tx", None) and getattr(s.tx, "strRef", None):
                    titles.append(str(s.tx.strRef.f))
                else:
                    titles.append("")
            except Exception:
                titles.append("")
            # Formula for values
            f = None
            try:
                if getattr(s, "values", None) is not None:
                    f = getattr(s.values, "formula", None) or getattr(s.values, "f", None)
            except Exception:
                f = None
            formulas.append(str(f) if f is not None else "")
    except Exception:
        pass
    return titles, formulas


def audit_workbook(inp: Path, outdir: Path, monthly_path: Optional[Path]) -> None:
    outdir.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(filename=str(inp), keep_vba=True, data_only=False)

    # Write sheets
    with (outdir / "sheets.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["sheet", "chart_count", "max_row", "max_column"])
        for ws in wb.worksheets:
            try:
                charts = getattr(ws, "_charts", []) or []
                w.writerow([ws.title, len(charts), ws.max_row, ws.max_column])
            except Exception:
                w.writerow([ws.title, "", ws.max_row, ws.max_column])

    # Named ranges
    with (outdir / "names.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["name", "refers_to"])
        for nm in wb.defined_names.definedName:
            w.writerow([nm.name, nm.attr_text])

    # Charts
    chart_rows: List[List[str]] = []
    for ws in wb.worksheets:
        charts = getattr(ws, "_charts", []) or []
        for ch in charts:
            title = _safe_title(ch)
            st, sf = _series_meta(ch)
            chart_rows.append([ws.title, title, "|".join(st), "|".join(sf)])

    with (outdir / "charts.csv").open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow(["sheet", "title", "series_titles", "series_formulas", "flag_low_signal", "reason"])

        monthly: Optional[pd.DataFrame] = None
        if monthly_path and monthly_path.exists():
            try:
                monthly = pd.read_csv(monthly_path, index_col=0, parse_dates=[0])
            except Exception:
                monthly = None

        for sheet, title, st_str, sf_str in chart_rows:
            flag = False
            reason = ""
            if monthly is not None:
                candidates = []
                # try to match by series titles or title tokens
                for token in (st_str.split("|") + title.replace("-", " ").split()):
                    tok = token.strip().upper()
                    if tok in monthly.columns:
                        candidates.append(tok)
                # compute 5y vol/coverage
                if candidates:
                    s = pd.to_numeric(monthly[candidates[0]], errors="coerce")
                    last5 = s.iloc[-60:]
                    cov = last5.notna().mean()
                    vol = last5.std(ddof=0)
                    if cov < 0.8:
                        flag = True
                        reason = f"coverage<80% ({cov:.1%})"
                    elif float(vol or 0.0) < 1e-6:
                        flag = True
                        reason = "vol≈0 over 5y"
            w.writerow([sheet, title, st_str, sf_str, "YES" if flag else "NO", reason])


def main():
    p = argparse.ArgumentParser(description="Audit Excel workbook charts and ranges")
    p.add_argument("--in", dest="inp", required=True, help="Input .xlsm path")
    p.add_argument("--out", dest="outdir", default="Output/audit", help="Output directory for CSVs")
    p.add_argument("--monthly", dest="monthly", help="Optional monthly dataframe CSV")
    args = p.parse_args()

    audit_workbook(Path(args.inp), Path(args.outdir), Path(args.monthly) if args.monthly else None)


if __name__ == "__main__":
    main()


