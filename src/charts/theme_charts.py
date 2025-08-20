from __future__ import annotations

from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from openpyxl.worksheet.worksheet import Worksheet

from .helpers import (
    breadth_z,
    credit_spread,
    insert_image,
    make_move,
    choose_first,
    find_series,
    normalize_series,
    theme_composite,
    theme_composite_from_cfg,
    rolling_percentile_cone,
    save_fig,
    safe_cols,
    to_periodic_rf,
    yoy,
    zscore,
)
from .helpers_series import align_monthly, ensure_series, ensure_series_by_key, aliases_for


OUTDIR = Path("Output/excel_charts")


def _resolve_sheet(book_or_ws, sheet_name: str) -> Worksheet:
    # If a worksheet is passed directly, use it
    try:
        from openpyxl.worksheet.worksheet import Worksheet as _WS
        if isinstance(book_or_ws, _WS):
            return book_or_ws
    except Exception:
        pass
    # Otherwise treat as workbook
    try:
        ws = book_or_ws[sheet_name]
    except Exception:
        ws = book_or_ws.create_sheet(sheet_name)
    return ws


def _note(ws: Worksheet, cell: str, msg: str) -> None:
    try:
        ws[cell].value = msg
    except Exception:
        pass


def build_growth_sheet(df: pd.DataFrame, book, cfg, sheet_name: str = "Growth", theme_df: Optional[pd.DataFrame] = None) -> None:
    if df is None or df.empty:
        return
    ws = _resolve_sheet(book, sheet_name)

    # G1: Composite + 5y cone
    comp_col = next((c for c in ["F_Growth", "Growth_Composite"] if c in df.columns), None)
    s = None
    if comp_col is not None:
        s = pd.to_numeric(df[comp_col], errors="coerce").dropna()
    else:
        s = theme_composite_from_cfg(df, cfg, "growth")
    # Fallback to ThemeComposite from the sheet's theme table if available
    if (s is None or len(s) == 0) and theme_df is not None and "ThemeComposite" in theme_df.columns:
        try:
            s = pd.to_numeric(theme_df["ThemeComposite"], errors="coerce").dropna()
        except Exception:
            s = None
    if s is not None and len(s) > 0:
        if len(s) > 0:
            # last 5 years when long
            s_plot = s.tail(60) if len(s) > 60 else s
            cone = rolling_percentile_cone(s, window=60).reindex(s_plot.index)
            fig, ax = plt.subplots(figsize=(7.0, 3.5))
            try:
                ax.fill_between(cone.index, cone["p10"], cone["p90"], alpha=0.2, label="5y cone")
                ax.plot(cone.index, cone["p50"], linewidth=1.2, alpha=0.8, label="median")
            except Exception:
                pass
            ax.plot(s_plot.index, s_plot.values, linewidth=1.8, label=comp_col)
            ax.set_title("Growth Composite — 5-Year Percentile Cone")
            ax.set_xlabel("Date")
            ax.grid(True, alpha=0.25)
            ax.legend(loc="best")
            p = save_fig(fig, OUTDIR, "growth_cone.png")
            insert_image(ws, p, "A32", max_width=440, max_height=260)
    else:
        _note(ws, "J2", "Growth composite not available — cone chart skipped")

    # G2: Breadth of z-scores (components)
    comp_list = getattr(cfg, "GROUPS", {}).get("growth", []) or getattr(cfg, "THEME_GROUPS", {}).get("growth", [])
    z_cols = [c for c in comp_list if f"{c}_ZScore" in df.columns]
    if not z_cols:
        # fallback: any *_ZScore present with growth-like naming
        z_cols = [c for c in df.columns if isinstance(c, str) and c.endswith("_ZScore") and "EMP" not in c.upper()]
    if z_cols:
        zdf = df[z_cols].apply(pd.to_numeric, errors="coerce")
        br = breadth_z(zdf).tail(180)
        if not br.empty:
            fig, ax = plt.subplots(figsize=(7.0, 3.5))
            br["pct_pos"].plot(ax=ax, label=">0")
            br["pct_pos1"].plot(ax=ax, label=">+1")
            ax.set_ylim(0, 100)
            ax.set_title("Growth Breadth (Diffusion)")
            ax.set_ylabel("Percent")
            ax.set_xlabel("Date")
            ax.grid(True, alpha=0.25)
            ax.legend(loc="best")
            p = save_fig(fig, OUTDIR, "growth_breadth.png")
            insert_image(ws, p, "J32", max_width=440, max_height=260)
    else:
        _note(ws, "J22", "Growth component z-scores not available — breadth chart skipped")

    # G3: INDPRO YoY vs PAYEMS YoY (with diff)
    a = df.get("INDPRO_YoY")
    b = df.get("PAYEMS_YoY")
    if a is None and "INDPRO" in df.columns:
        a = yoy(df["INDPRO"]) * 100
    if b is None and "PAYEMS" in df.columns:
        b = yoy(df["PAYEMS"]) * 100
    if a is not None and b is not None:
        idx = a.dropna().index.intersection(b.dropna().index)
        a2 = pd.to_numeric(a, errors="coerce").reindex(idx).dropna()
        b2 = pd.to_numeric(b, errors="coerce").reindex(idx).dropna()
        if not a2.empty and not b2.empty:
            diff = a2 - b2
            fig, ax = plt.subplots(figsize=(7.0, 3.5))
            a2.tail(240).plot(ax=ax, label="INDPRO YoY")
            b2.tail(240).plot(ax=ax, label="PAYEMS YoY")
            ax2 = ax.twinx()
            diff.tail(240).plot(ax=ax2, color="#777777", linewidth=1.2, label="Diff (L2R)")
            ax.set_title("Production vs Employment YoY (with Divergence)")
            ax.set_ylabel("% YoY")
            ax.set_xlabel("Date")
            ax.grid(True, alpha=0.25)
            try:
                h1, l1 = ax.get_legend_handles_labels()
                h2, l2 = ax2.get_legend_handles_labels()
                ax.legend(h1 + h2, l1 + l2, loc="best")
            except Exception:
                ax.legend(loc="best")
            p = save_fig(fig, OUTDIR, "growth_divergence.png")
            insert_image(ws, p, "A62", max_width=440, max_height=260)
    else:
        _note(ws, "J42", "INDPRO/PAYEMS YoY not available — divergence chart skipped")


def build_housing_sheet(df: pd.DataFrame, book, cfg, sheet_name: str = "Housing") -> None:
    if df is None or df.empty:
        return
    ws = _resolve_sheet(book, sheet_name)

    # H1: Housing Affordability Ratio + Cone
    # Inputs via ensure_series
    s_rate = ensure_series_by_key(df, "MORTGAGE_RATE")
    s_price = ensure_series_by_key(df, "HOUSE_PRICE")
    s_income = ensure_series_by_key(df, "HOUSE_INCOME")
    if s_rate is None or s_price is None:
        _note(ws, "A42", f"Missing MORTGAGE_RATE or HOUSE_PRICE — tried {aliases_for('MORTGAGE_RATE')} / {aliases_for('HOUSE_PRICE')}")
    else:
        # scale rate to fraction if likely in percent
        s_rate = pd.to_numeric(s_rate, errors="coerce")
        if s_rate.abs().median(skipna=True) > 1.0:
            s_rate = s_rate / 100.0
        s_price_n = normalize_series(s_price)
        s_income_n = normalize_series(s_income) if s_income is not None else None
        aligned, ok = align_monthly(s_rate, s_price_n, s_income_n, ffill_limit=2)
        if not ok:
            _note(ws, "A42", "Mortgage rate and/or house price index missing or insufficient overlap — chart skipped")
        else:
            r, pz, inc = aligned
            if inc is None or inc.isna().all():
                affordability = (r * pz)
                title_suffix = " (no income proxy)"
            else:
                affordability = (r * pz) / inc
                title_suffix = ""
            aff = affordability.dropna().tail(180)
            cone = rolling_percentile_cone(affordability.dropna(), window=60).reindex(aff.index)
            fig, ax = plt.subplots(figsize=(7.0, 3.5))
            try:
                ax.fill_between(cone.index, cone["p10"], cone["p90"], alpha=0.2, label="5y cone")
            except Exception:
                pass
            ax.plot(aff.index, aff.values, linewidth=1.8, label="Affordability")
            ax.set_title("Housing Affordability Ratio + Cone" + title_suffix)
            ax.set_xlabel("Date")
            ax.grid(True, alpha=0.25)
            ax.legend(loc="best")
            p = save_fig(fig, OUTDIR, "housing_afford.png")
            insert_image(ws, p, "A62", max_width=440, max_height=260)

    # H2: Permits vs Starts (Lead/Lag)
    s_permit = ensure_series_by_key(df, "PERMITS")
    s_starts = ensure_series_by_key(df, "STARTS")
    if s_permit is None or s_starts is None:
        _note(ws, "A22", f"Missing PERMITS or STARTS — tried {aliases_for('PERMITS')} / {aliases_for('STARTS')}")
    else:
        aligned, ok = align_monthly(s_permit, s_starts, ffill_limit=6)
        if ok:
            a_permit, a_starts = aligned
            sub = pd.DataFrame({"PERMIT": a_permit, "HOUST": a_starts}).tail(240)
            fig, ax = plt.subplots(figsize=(7.0, 3.5))
            sub.plot(ax=ax, linewidth=1.5)
            try:
                lead3 = a_permit.shift(-3)
                lead3.tail(240).plot(ax=ax, linewidth=1.0, alpha=0.6, linestyle="--", label="Permits +3m")
            except Exception:
                pass
            ax.set_title("Permits vs Starts (Lead/Lag)")
            ax.set_xlabel("Date")
            ax.grid(True, alpha=0.25)
            ax.legend(loc="best")
            p = save_fig(fig, OUTDIR, "housing_permits_starts.png")
            insert_image(ws, p, "J32", max_width=440, max_height=260)
        else:
            # Fallback: align on union with extended ffill; proceed if we have >= 12 points
            try:
                a, b = align_monthly(s_permit, s_starts, ffill_limit=6)[0]
                idx = a.index.union(b.index)
                a2 = a.reindex(idx)
                b2 = b.reindex(idx)
                sub = pd.DataFrame({"PERMIT": a2, "HOUST": b2}).dropna(how="all").tail(240)
                if len(sub.dropna(how="any")) >= 12:
                    fig, ax = plt.subplots(figsize=(7.0, 3.5))
                    sub.plot(ax=ax, linewidth=1.5)
                    try:
                        lead3 = a2.shift(-3)
                        lead3.tail(240).plot(ax=ax, linewidth=1.0, alpha=0.6, linestyle="--", label="Permits +3m")
                    except Exception:
                        pass
                    ax.set_title("Permits vs Starts (Lead/Lag)")
                    ax.set_xlabel("Date")
                    ax.grid(True, alpha=0.25)
                    ax.legend(loc="best")
                    p = save_fig(fig, OUTDIR, "housing_permits_starts.png")
                    insert_image(ws, p, "J32", max_width=440, max_height=260)
                else:
                    _note(ws, "J32", "PERMIT and HOUST have insufficient overlap — chart skipped")
            except Exception:
                _note(ws, "J32", "PERMIT and HOUST have no overlapping non-empty area — chart skipped")

    # H3: Builder Sentiment vs Permits
    s_sent = ensure_series_by_key(df, "NAHB")
    if s_sent is None:
        s_sent = df.get("F_Housing")
    s_perm = ensure_series_by_key(df, "PERMITS")
    if s_sent is None or s_perm is None:
        _note(ws, "A2", f"NAHB (or F_Housing) and/or PERMIT(S) missing — tried {aliases_for('NAHB')} / {aliases_for('PERMITS')}")
    else:
        aligned, ok = align_monthly(s_sent, s_perm, ffill_limit=6)
        if ok:
            a_sent, a_perm = aligned
            a_sent = a_sent.tail(240)
            a_perm = a_perm.tail(240)
            fig, ax = plt.subplots(figsize=(7.0, 3.5))
            a_sent.plot(ax=ax, label="NAHB" if "NAHB" in df.columns else "F_Housing", linewidth=1.8)
            ax2 = ax.twinx()
            a_perm.plot(ax=ax2, color="#777777", linewidth=1.2, label="Permits")
            ax.set_title("Builder Sentiment vs Permits")
            ax.set_xlabel("Date")
            ax.grid(True, alpha=0.25)
            try:
                h1, l1 = ax.get_legend_handles_labels()
                h2, l2 = ax2.get_legend_handles_labels()
                ax.legend(h1 + h2, l1 + l2, loc="best")
            except Exception:
                ax.legend(loc="best")
            p = save_fig(fig, OUTDIR, "housing_sentiment.png")
            insert_image(ws, p, "A32", max_width=440, max_height=260)
        else:
            # Fallback: align on union with extended ffill; proceed if we have >= 12 points
            try:
                a, b = align_monthly(s_sent, s_perm, ffill_limit=6)[0]
                idx = a.index.union(b.index)
                a_sent = a.reindex(idx).tail(240)
                a_perm = b.reindex(idx).tail(240)
                if a_sent.dropna().size >= 12 and a_perm.dropna().size >= 12:
                    fig, ax = plt.subplots(figsize=(7.0, 3.5))
                    a_sent.plot(ax=ax, label="NAHB" if "NAHB" in df.columns else "F_Housing", linewidth=1.8)
                    ax2 = ax.twinx()
                    a_perm.plot(ax=ax2, color="#777777", linewidth=1.2, label="Permits")
                    ax.set_title("Builder Sentiment vs Permits")
                    ax.set_xlabel("Date")
                    ax.grid(True, alpha=0.25)
                    try:
                        h1, l1 = ax.get_legend_handles_labels()
                        h2, l2 = ax2.get_legend_handles_labels()
                        ax.legend(h1 + h2, l1 + l2, loc="best")
                    except Exception:
                        ax.legend(loc="best")
                    p = save_fig(fig, OUTDIR, "housing_sentiment.png")
                    insert_image(ws, p, "A32", max_width=440, max_height=260)
                else:
                    _note(ws, "A32", "NAHB/F_Housing and PERMIT have insufficient overlap — chart skipped")
            except Exception:
                _note(ws, "A32", "NAHB/F_Housing and PERMIT have no overlapping non-empty area — chart skipped")


def build_inflation_sheet(df: pd.DataFrame, book, cfg, sheet_name: str = "Inflation", theme_df: Optional[pd.DataFrame] = None) -> None:
    if df is None or df.empty:
        return
    ws = _resolve_sheet(book, sheet_name)

    # I1: Composite + 5y cone
    comp_col = next((c for c in ["F_Inflation", "Inflation_Composite"] if c in df.columns), None)
    s = None
    if comp_col is not None:
        s = pd.to_numeric(df[comp_col], errors="coerce").dropna()
    else:
        s = theme_composite_from_cfg(df, cfg, "inflation")
    if (s is None or len(s) == 0) and theme_df is not None and "ThemeComposite" in theme_df.columns:
        try:
            s = pd.to_numeric(theme_df["ThemeComposite"], errors="coerce").dropna()
        except Exception:
            s = None
    if s is not None and len(s) > 0:
        s_plot = s.tail(60) if len(s) > 60 else s
        cone = rolling_percentile_cone(s, window=60).reindex(s_plot.index)
        fig, ax = plt.subplots(figsize=(7.0, 3.5))
        try:
            ax.fill_between(cone.index, cone["p10"], cone["p90"], alpha=0.2, label="5y cone")
            ax.plot(cone.index, cone["p50"], linewidth=1.2, alpha=0.8, label="median")
        except Exception:
            pass
        ax.plot(s_plot.index, s_plot.values, linewidth=1.8, label=comp_col)
        ax.set_title("Inflation Composite — 5-Year Percentile Cone")
        ax.set_xlabel("Date")
        ax.grid(True, alpha=0.25)
        ax.legend(loc="best")
        p = save_fig(fig, OUTDIR, "inflation_cone.png")
        insert_image(ws, p, "A32", max_width=440, max_height=260)
    else:
        _note(ws, "J2", "Inflation composite not available — cone chart skipped")

    # I2: Core − Headline spread via ensure_series and alignment
    head = ensure_series_by_key(df, "CPI_YOY")
    core = ensure_series_by_key(df, "CORECPI_YOY")
    if head is None or core is None:
        _note(ws, "J22", f"Core or headline inflation missing — tried {aliases_for('CORECPI_YOY')} / {aliases_for('CPI_YOY')}")
    else:
        aligned, ok = align_monthly(core, head, ffill_limit=6)
        if ok:
            a_core, a_head = aligned
            s = (a_core - a_head).dropna().tail(180)
            if not s.empty:
                fig, ax = plt.subplots(figsize=(7.0, 3.5))
                s.plot(ax=ax, linewidth=1.8, label="Core - Headline")
                ax.axhline(0.0, color="#888888", linewidth=1.0)
                ax.set_title("Core − Headline CPI Spread")
                ax.set_xlabel("Date")
                ax.grid(True, alpha=0.25)
                ax.legend(loc="best")
                try:
                    ws["J22"].value = None
                except Exception:
                    pass
                p = save_fig(fig, OUTDIR, "inflation_core_spread.png")
                insert_image(ws, p, "J32", max_width=440, max_height=260)
        else:
            _note(ws, "J22", "Core or headline inflation missing — spread chart skipped")

    # I3: Liquidity pulse (M2 YoY & real policy rate); plot both curve and real rate when available
    m2 = df.get("M2_YoY")
    if m2 is None:
        m2_src = choose_first(df, ["M2SL", "M2", "M2NSA"]) 
        if m2_src and m2_src in df.columns:
            m2 = yoy(df[m2_src]) * 100
    real_rate = None
    if "FEDFUNDS" in df.columns:
        cpi_yoy = df.get("CPI_YoY")
        if cpi_yoy is None and "CPIAUCSL" in df.columns:
            cpi_yoy = yoy(df["CPIAUCSL"]) * 100
        if cpi_yoy is not None:
            real_rate = pd.to_numeric(df["FEDFUNDS"], errors="coerce") - pd.to_numeric(cpi_yoy, errors="coerce")
    curve = df.get("T10Y3M")
    if m2 is not None and (real_rate is not None or curve is not None):
        fig, ax = plt.subplots(figsize=(7.0, 3.5))
        pd.to_numeric(m2, errors="coerce").tail(180).plot(ax=ax, label="M2 YoY")
        ax2 = ax.twinx()
        plotted_any = False
        if real_rate is not None:
            pd.to_numeric(real_rate, errors="coerce").tail(180).plot(ax=ax2, color="#444444", linewidth=1.2, label="Real Policy Rate")
            plotted_any = True
        if curve is not None:
            pd.to_numeric(curve, errors="coerce").tail(180).plot(ax=ax2, color="#777777", linewidth=1.2, linestyle="--", label="Curve (10y-3m)")
            plotted_any = True
        if plotted_any:
            ax.set_title("Liquidity Pulse")
            ax.set_xlabel("Date")
            ax.grid(True, alpha=0.25)
            try:
                h1, l1 = ax.get_legend_handles_labels()
                h2, l2 = ax2.get_legend_handles_labels()
                ax.legend(h1 + h2, l1 + l2, loc="best")
            except Exception:
                ax.legend(loc="best")
            p = save_fig(fig, OUTDIR, "inflation_liquidity.png")
            insert_image(ws, p, "A62", max_width=440, max_height=260)
        else:
            _note(ws, "J42", "No valid real rate or curve after alignment — liquidity chart skipped")
    else:
        _note(ws, "J42", "M2 YoY and/or real rate/curve missing — liquidity chart skipped")


def build_credit_sheet(df: pd.DataFrame, book, cfg, sheet_name: str = "Credit") -> None:
    if df is None or df.empty:
        return
    ws = _resolve_sheet(book, sheet_name)

    vix_name = choose_first(df, ["VIXCLS", "VIX", "VIX_SA"]) 
    vix = pd.to_numeric(df[vix_name], errors="coerce") if vix_name else None
    move = make_move(df)
    spr = credit_spread(df)
    parts = [x for x in [vix, move, spr] if x is not None]
    if not parts:
        _note(ws, "A2", "Credit stress inputs missing — stress chart skipped")
        return
    # align
    idx = parts[0].index
    for p in parts[1:]:
        idx = idx.intersection(p.index)
    parts = [p.reindex(idx) for p in parts]
    zs = [zscore(p) for p in parts]
    stress_z = pd.DataFrame(zs).T.mean(axis=1)

    # C1: mini line (last 60m)
    s60 = stress_z.dropna().tail(60)
    if not s60.empty:
        fig, ax = plt.subplots(figsize=(7.0, 3.5))
        s60.plot(ax=ax, label="Stress z")
        ax.axhline(0.0, color="#888888", linewidth=1.0)
        ax.set_title("Credit/Risk Stress (z)")
        ax.set_xlabel("Date")
        ax.grid(True, alpha=0.25)
        ax.legend(loc="best")
        p = save_fig(fig, OUTDIR, "credit_stress_line.png")
        insert_image(ws, p, "A32", max_width=440, max_height=260)

    # Histogram bars: latest vs p10/p50/p90
    s_all = stress_z.dropna()
    if not s_all.empty:
        latest = float(s_all.iloc[-1])
        p10, p50, p90 = np.nanpercentile(s_all.values, [10, 50, 90])
        data = pd.Series({"p10": p10, "p50": p50, "p90": p90, "latest": latest})
        fig, ax = plt.subplots(figsize=(7.0, 3.5))
        data[["p10", "p50", "p90"]].plot(kind="bar", ax=ax)
        ax.axhline(latest, color="#cc3333", linewidth=2.0, label="latest")
        ax.set_title("Stress distribution: p10/p50/p90 vs latest")
        ax.grid(True, axis="y", alpha=0.25)
        ax.legend(loc="best")
        p = save_fig(fig, OUTDIR, "credit_stress_bar.png")
        insert_image(ws, p, "J32", max_width=440, max_height=260)


