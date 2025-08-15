"""
Headless Excel dashboard builder using openpyxl and matplotlib.
Rebuilds visuals in-place and supports fallbacks to processed CSVs if the
template lacks pre-populated tables.
"""
from __future__ import annotations

from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
try:
    from src.charts.theme_charts import (
        build_growth_sheet,
        build_housing_sheet,
        build_inflation_sheet,
        build_credit_sheet,
    )
except Exception:  # pragma: no cover
    build_growth_sheet = build_housing_sheet = build_inflation_sheet = build_credit_sheet = None  # type: ignore
try:
    from config import config as CHART_CFG  # type: ignore
except Exception:  # pragma: no cover
    CHART_CFG = object()
try:
    # Optional: portfolio/dashboard configuration for display
    from config.config import REGIME_WINDOW_YEARS, REBAL_FREQ, TRANSACTION_COST
except Exception:  # pragma: no cover
    REGIME_WINDOW_YEARS, REBAL_FREQ, TRANSACTION_COST = 10, 'M', 0.0005

# Optional heavy dependency; guarded import
try:
    from src.models.portfolio import PortfolioConstructor, OptimizationMethod
except Exception:  # pragma: no cover - optional import for dashboard extras
    PortfolioConstructor = None  # type: ignore
    OptimizationMethod = None  # type: ignore

THEME_SHEETS: List[str] = [
    "Growth & Labour",
    "Inflation & Liquidity",
    "Credit & Risk",
    "Housing",
    "FX & Commodities",
]


THEME_COLOR: Dict[str, str] = {
    "Growth & Labour": "#1f77b4",
    "Inflation & Liquidity": "#d62728",
    "Credit & Risk": "#ff7f0e",
    "Housing": "#2ca02c",
    "FX & Commodities": "#9467bd",
}


def load_workbook_safe(path: str | Path):
    path_str = str(path)
    keep_vba = path_str.lower().endswith(".xlsm")
    return load_workbook(filename=path_str, keep_vba=keep_vba)


def _save_plot(fig, outdir: Path, filename: str) -> Path:
    outdir.mkdir(parents=True, exist_ok=True)
    out = outdir / filename
    fig.savefig(out, dpi=140, bbox_inches="tight")
    plt.close(fig)
    return out


def insert_image(ws: Worksheet, image_path: Path, anchor_cell: str, *, max_width: int | None = None, max_height: int | None = None) -> None:
    """Insert an image with optional bounding-box scaling to avoid overlaps.

    Sizes are in pixels; aspect ratio is preserved when any bound is set.
    """
    try:
        img = XLImage(str(image_path))
        try:
            orig_w, orig_h = int(img.width), int(img.height)
        except Exception:
            orig_w = orig_h = None
        if orig_w and orig_h and (max_width or max_height):
            scale_w = (max_width / orig_w) if max_width else 1.0
            scale_h = (max_height / orig_h) if max_height else 1.0
            scale = min(scale_w, scale_h)
            if scale < 1.0:
                img.width = int(orig_w * scale)
                img.height = int(orig_h * scale)
        ws.add_image(img, anchor_cell)
    except Exception:
        pass


def _read_table_from_sheet(ws: Worksheet) -> Optional[pd.DataFrame]:
    max_rows, max_cols = 200, 40
    values = [[ws.cell(r, c).value for c in range(1, max_cols + 1)] for r in range(1, max_rows + 1)]
    header_row = None
    for r in range(len(values)):
        row = values[r]
        if row and any((isinstance(v, str) and "themecomposite" in v.lower()) for v in row if v):
            header_row = r
            break
    if header_row is None:
        return None
    headers = values[header_row]
    last_col = max(i for i, v in enumerate(headers, start=1) if v is not None)
    data_rows: List[List[object]] = []
    for r in range(header_row + 1, len(values)):
        row = values[r][: last_col]
        if all(v is None for v in row):
            break
        data_rows.append(row)
    if not data_rows:
        return None
    headers = [h if h is not None else "Date" for h in headers[: last_col]]
    df = pd.DataFrame(data_rows, columns=headers)
    try:
        df.iloc[:, 0] = pd.to_datetime(df.iloc[:, 0])
        df = df.set_index(df.columns[0])
    except Exception:
        pass
    return df


def load_theme_df(wb, sheetname: str) -> Optional[pd.DataFrame]:
    if sheetname not in wb.sheetnames:
        return None
    ws: Worksheet = wb[sheetname]
    df = _read_table_from_sheet(ws)
    if df is None:
        return None
    needed = {"ThemeComposite", "ThemeComposite_3M_MA", "ThemeComposite_YoY"}
    if not needed.intersection(set(df.columns)):
        return None
    return df


def compute_metrics(df: pd.DataFrame):
    comp = df["ThemeComposite"].astype(float)
    latest = comp.iloc[-1]
    chg_3m = comp.diff(3).iloc[-1] if len(comp) >= 4 else np.nan
    yoy = df["ThemeComposite_YoY"].astype(float).iloc[-1] if "ThemeComposite_YoY" in df.columns else np.nan
    ranks = comp.rank(method="average")
    percentile = 100 * ranks / ranks.max()
    window = 36 if len(comp) >= 36 else 24 if len(comp) >= 24 else max(6, len(comp))
    roll_mean = comp.rolling(window, min_periods=max(3, window // 3)).mean()
    roll_std = comp.rolling(window, min_periods=max(3, window // 3)).std(ddof=0)
    zscore = (comp - roll_mean) / (roll_std.replace(0, np.nan))
    roll_vol_12m = comp.rolling(12, min_periods=6).std(ddof=0)
    slope_6m = np.nan
    if len(comp) >= 6:
        y = comp.iloc[-6:]
        x = np.arange(len(y))
        coeffs = np.polyfit(x, y.values, 1)
        slope_6m = float(coeffs[0])
    return latest, chg_3m, yoy, percentile, zscore, roll_vol_12m, slope_6m


def load_data_dump_df(wb) -> Optional[pd.DataFrame]:
    """Read the 'Data Dump' sheet as a DataFrame.

    Assumes headers on first row, index in first column (dates).
    """
    if "Data Dump" not in wb.sheetnames:
        return None
    ws: Worksheet = wb["Data Dump"]
    values = list(ws.values)
    if not values:
        return None
    headers = [h if h is not None else f"col_{i}" for i, h in enumerate(values[0])]
    rows = values[1:]
    if not rows:
        return None
    df = pd.DataFrame(rows, columns=headers)
    try:
        df.iloc[:, 0] = pd.to_datetime(df.iloc[:, 0])
        df = df.set_index(df.columns[0])
    except Exception:
        pass
    return df


def render_plot_trend(theme: str, df: pd.DataFrame, outdir: Path) -> Optional[Path]:
    try:
        color = THEME_COLOR.get(theme, "#1f77b4")
        fig, ax = plt.subplots(figsize=(7.4, 3.8))
        # Show full history when short; otherwise show the most recent ~15 years (180 months)
        base = df
        try:
            if isinstance(df.index, pd.DatetimeIndex) and len(df) > 180:
                base = df.tail(180)
        except Exception:
            base = df
        base["ThemeComposite"].plot(ax=ax, color=color, linewidth=1.8, label="Composite")
        if "ThemeComposite_3M_MA" in base.columns:
            base["ThemeComposite_3M_MA"].plot(ax=ax, color="#999999", linewidth=1.5, linestyle="--", label="3M MA")
        ax.set_title(f"{theme} Composite — Trend")
        ax.set_ylabel("Index")
        ax.set_xlabel("Date")
        ax.grid(True, alpha=0.25)
        ax.legend(loc="best")
        return _save_plot(fig, outdir, f"{theme.replace(' & ', '_').replace(' ', '_').lower()}_trend.png")
    except Exception:
        return None


def render_theme_signature(theme: str, monthly: pd.DataFrame, outdir: Path) -> Optional[Path]:
    """Render a theme-specific signature chart. Skip gracefully if data missing."""
    try:
        fig, ax = plt.subplots(figsize=(7.4, 3.8))
        t = theme
        m = monthly
        used_ax2 = False
        if t == "Growth & Labour":
            # Business cycle clock: GDP_YoY vs UNRATE, colored by date order
            if all(c in m.columns for c in ("GDP_YoY", "UNRATE")):
                xy = m[["GDP_YoY", "UNRATE"]].dropna()
                if xy.empty:
                    raise ValueError
                colors = np.linspace(0, 1, len(xy))
                sc = ax.scatter(xy["GDP_YoY"], xy["UNRATE"], c=colors, cmap="viridis", s=12)
                ax.set_xlabel("GDP YoY (%)")
                ax.set_ylabel("Unemployment Rate (%)")
                ax.set_title("Business Cycle Clock")
            else:
                ax.text(0.1, 0.5, "Data not available", transform=ax.transAxes)
                ax.set_axis_off()

        elif t == "Inflation & Liquidity":
            # CPI contributions (if component series exist) and real policy rate line
            components = [c for c in ("CPIAPPSL", "CPITRNSL", "CPIMEDSL", "CUSR0000SAC") if c in m.columns]
            plotted = False
            if components:
                comp_yoy = pd.DataFrame({c: pd.to_numeric(m[c], errors="coerce").pct_change(12) * 100 for c in components}).dropna()
                if not comp_yoy.empty:
                    comp_yoy.tail(180).plot.area(ax=ax, stacked=True, linewidth=0)
                    plotted = True
            if "FEDFUNDS" in m.columns and "CPI_YoY" in m.columns:
                rr = (pd.to_numeric(m["FEDFUNDS"], errors="coerce") - pd.to_numeric(m["CPI_YoY"], errors="coerce")).dropna()
                if not plotted:
                    rr.plot(ax=ax, color="#444444", linewidth=1.5, label="Real Policy Rate")
                else:
                    ax2 = ax.twinx()
                    rr.tail(180).plot(ax=ax2, color="#444444", linewidth=1.5, label="Real Policy Rate")
                    ax2.set_ylabel("Real Rate (%)")
                    ax2.set_ylim(rr.tail(180).min()*1.1, rr.tail(180).max()*1.1)
                    used_ax2 = True
                plotted = True
            if not plotted:
                ax.text(0.1, 0.5, "Data not available", transform=ax.transAxes)
                ax.set_axis_off()
            else:
                ax.set_title("CPI Contributions & Real Policy Rate")
                ax.set_xlabel("Date")

        elif t == "Credit & Risk":
            # IG spread (BAA-AAA) and NFCI stress if available
            plotted = False
            if all(c in m.columns for c in ("BAA", "AAA")):
                ig = (pd.to_numeric(m["BAA"], errors="coerce") - pd.to_numeric(m["AAA"], errors="coerce")).dropna()
                ig.tail(180).plot(ax=ax, color="#cc5500", linewidth=1.5, label="IG Spread (BAA-AAA)")
                ax.set_ylabel("Spread (pp)")
                plotted = True
            if "NFCI" in m.columns:
                ax2 = ax.twinx() if plotted else ax
                pd.to_numeric(m["NFCI"], errors="coerce").tail(180).plot(ax=ax2, color="#5555aa", linewidth=1.2, label="NFCI")
                if ax2 is not ax:
                    ax2.set_ylabel("NFCI")
                    used_ax2 = True
                plotted = True
            if not plotted:
                ax.text(0.1, 0.5, "Data not available", transform=ax.transAxes)
                ax.set_axis_off()
            else:
                ax.set_title("Credit Spreads & Financial Stress")
                ax.grid(True, axis="y", alpha=0.2)
                ax.set_xlabel("Date")

        elif t == "Housing":
            # Permits vs Starts; mortgage rate if available
            plotted = False
            cols = [c for c in ("PERMIT", "HOUST") if c in m.columns]
            if cols:
                pd.DataFrame({c: pd.to_numeric(m[c], errors="coerce") for c in cols}).tail(240).plot(ax=ax, linewidth=1.5)
                ax.set_ylabel("Index")
                plotted = True
            if "MORTGAGE30US" in m.columns:
                ax2 = ax.twinx()
                pd.to_numeric(m["MORTGAGE30US"], errors="coerce").tail(240).plot(ax=ax2, color="#777777", linewidth=1.2, label="Mortgage 30Y")
                ax2.set_ylabel("Mortgage 30Y (%)")
                used_ax2 = True
                plotted = True
            if not plotted:
                ax.text(0.1, 0.5, "Data not available", transform=ax.transAxes)
                ax.set_axis_off()
            else:
                ax.set_title("Housing Pipeline & Mortgage Rate")
                ax.set_xlabel("Date")

        elif t == "FX & Commodities":
            # Energy proxy (WTI), Dollar index (TWEX), optional gold if present
            plotted = False
            if "DCOILWTICO" in m.columns:
                pd.to_numeric(m["DCOILWTICO"], errors="coerce").tail(240).plot(ax=ax, color="#8c564b", linewidth=1.5, label="WTI Oil")
                ax.set_ylabel("WTI (USD)")
                plotted = True
            if "TWEXAFEGSMTH" in m.columns:
                ax2 = ax.twinx() if plotted else ax
                pd.to_numeric(m["TWEXAFEGSMTH"], errors="coerce").tail(240).plot(ax=ax2, color="#1f77b4", linewidth=1.2, label="Dollar Index")
                if ax2 is not ax:
                    ax2.set_ylabel("Dollar Index")
                    used_ax2 = True
                plotted = True
            if not plotted:
                ax.text(0.1, 0.5, "Data not available", transform=ax.transAxes)
                ax.set_axis_off()
            else:
                ax.set_title("Energy & Dollar Index")
                ax.set_xlabel("Date")

        else:
            ax.text(0.1, 0.5, "Data not available", transform=ax.transAxes)
            ax.set_axis_off()

        # Unify legends across primary and twin axes when present
        try:
            lines, labels = ax.get_legend_handles_labels()
            if used_ax2:
                ax2 = ax.twinx()
                # immediately remove to avoid double axis; only for legend harvest
                ax.figure.delaxes(ax2)
            # Instead, fetch from existing twin if it exists
        except Exception:
            pass
        try:
            # Safer legend merge: query all axes in the figure
            handles, labels = [], []
            for a in fig.axes:
                h, l = a.get_legend_handles_labels()
                handles.extend(h); labels.extend(l)
            if labels:
                fig.legend(handles, labels, loc="upper right")
        except Exception:
            pass
        return _save_plot(fig, outdir, f"{theme.replace(' & ', '_').replace(' ', '_').lower()}_signature.png")
    except Exception:
        return None


def build_theme_sheet(wb, sheetname: str, outdir: Path) -> None:
    if sheetname not in wb.sheetnames:
        return
    ws: Worksheet = wb[sheetname]
    # Remove any previously inserted images so new renders don't stack
    try:
        for img in list(getattr(ws, "_images", [])):
            ws._images.remove(img)
    except Exception:
        pass
    df = load_theme_df(wb, sheetname)
    if df is None or df.empty:
        return
    p1 = render_plot_trend(sheetname, df, outdir)
    if p1:
        insert_image(ws, p1, "A2", max_width=620, max_height=280)
    # Signature chart using Data Dump
    monthly = load_data_dump_df(wb)
    if monthly is not None and not monthly.empty:
        p2 = render_theme_signature(sheetname, monthly, outdir)
        if p2:
            insert_image(ws, p2, "A22", max_width=620, max_height=280)
    # Also place the new high-signal charts using the same pipeline to avoid
    # any writer/adapter discrepancies. We read the featured CSV if available,
    # falling back to the monthly Data Dump.
    try:
        if build_growth_sheet:
            # Load macro_df (featured preferred)
            macro_df = None
            try:
                featured_csv = Path("Data/processed/macro_data_featured.csv")
                if featured_csv.exists():
                    macro_df = pd.read_csv(featured_csv, index_col=0, parse_dates=[0])
            except Exception:
                macro_df = None
            if macro_df is None:
                macro_df = monthly
            if macro_df is not None and not macro_df.empty:
                if sheetname == "Growth & Labour" or sheetname == "Growth":
                    build_growth_sheet(macro_df, wb, CHART_CFG, sheet_name=sheetname)
                elif sheetname == "Inflation & Liquidity" or sheetname == "Inflation":
                    build_inflation_sheet(macro_df, wb, CHART_CFG, sheet_name=sheetname)
                elif sheetname == "Credit & Risk" or sheetname == "Credit":
                    build_credit_sheet(macro_df, wb, CHART_CFG, sheet_name=sheetname)
                elif sheetname == "Housing":
                    build_housing_sheet(macro_df, wb, CHART_CFG, sheet_name=sheetname)
    except Exception:
        pass


def _build_6040_benchmark(returns: pd.DataFrame) -> pd.Series | None:
    """Create a simple 60/40 stock/bond benchmark from available return columns.

    Heuristics: prefer 'SPX' for stocks and 'US10Y_NOTE_FUT' for bonds, with
    several fallbacks. Returns are expected as arithmetic monthly returns.
    """
    if returns is None or returns.empty:
        return None
    cols = list(returns.columns)
    def pick(preferences: list[str]) -> str | None:
        for p in preferences:
            if p in cols:
                return p
        # loose match
        for c in cols:
            lc = str(c).lower()
            if any(k in lc for k in [p.lower() for p in preferences]):
                return c
        return None
    stock_pref = ["SPX", "SP500", "^GSPC", "SPTR", "EQUITY", "STOCK"]
    bond_pref = ["US10Y_NOTE_FUT", "US30Y_BOND_FUT", "ZN", "ZB", "IEF", "AGG", "BOND"]
    stock_col = pick(stock_pref)
    bond_col = pick(bond_pref)
    if not stock_col or not bond_col:
        return None
    s = pd.to_numeric(returns[stock_col], errors="coerce").fillna(0.0)
    b = pd.to_numeric(returns[bond_col], errors="coerce").fillna(0.0)
    bench = 0.6 * s + 0.4 * b
    return bench


def build_dashboard(wb, outdir: Path) -> None:
    if "Dashboard" not in wb.sheetnames:
        return

    def _compute_theme_series_fallback() -> dict:
        featured_csv = Path("Data/processed/macro_data_featured.csv")
        if not featured_csv.exists():
            return {}
        try:
            m = pd.read_csv(featured_csv, index_col=0, parse_dates=[0])
        except Exception:
            return {}
        groups = {"Growth & Labour": [], "Inflation & Liquidity": [], "Credit & Risk": [], "Housing": [], "FX & Commodities": []}
        infl_kw = ("cpi", "infl", "ppi", "m2", "liqu")
        credit_kw = ("spread", "risk", "nfc", "baml", "move")
        housing_kw = ("houst", "housing", "case", "mort", "permit")
        fx_kw = ("fx", "usd", "eur", "oil", "gold", "btc", "commod", "dcoil")
        for col in m.columns:
            lc = str(col).lower()
            if lc.startswith(infl_kw) or any(k in lc for k in infl_kw):
                groups["Inflation & Liquidity"].append(col)
            elif any(k in lc for k in credit_kw):
                groups["Credit & Risk"].append(col)
            elif any(k in lc for k in housing_kw):
                groups["Housing"].append(col)
            elif any(k in lc for k in fx_kw):
                groups["FX & Commodities"].append(col)
            else:
                groups["Growth & Labour"].append(col)
        out = {}
        for theme, cols in groups.items():
            if not cols:
                continue
            sub = m[cols].dropna(how="all")
            if sub.empty:
                continue
            z = sub.apply(lambda s: (s - s.mean()) / (s.std(ddof=0) or 1.0))
            out[theme] = pd.to_numeric(z.mean(axis=1), errors="coerce")
        return out

    # Correlation heatmap
    series: Dict[str, pd.Series] = {}
    for theme in THEME_SHEETS:
        df = load_theme_df(wb, theme)
        if df is not None and not df.empty and "ThemeComposite" in df.columns:
            series[theme] = pd.to_numeric(df["ThemeComposite"], errors="coerce")
    if len(series) < 3:
        series = _compute_theme_series_fallback()
    if len(series) >= 3:
        dfc = pd.DataFrame(series).dropna(how="all")
        if isinstance(dfc.index, pd.DatetimeIndex) and len(dfc) > 0:
            cutoff = dfc.index.max() - pd.DateOffset(years=5)
            dfc = dfc[dfc.index >= cutoff]
        if not dfc.empty:
            try:
                ws: Worksheet = wb["Dashboard"]
                for img in list(getattr(ws, "_images", [])):
                    ws._images.remove(img)
            except Exception:
                pass
            try:
                corr = dfc.corr()
                fig, ax = plt.subplots(figsize=(7.4, 4.8))
                im = ax.imshow(corr, cmap="RdBu_r", vmin=-1, vmax=1)
                ax.set_xticks(range(len(corr.columns)))
                ax.set_xticklabels(corr.columns, rotation=45, ha="right")
                ax.set_yticks(range(len(corr.index)))
                ax.set_yticklabels(corr.index)
                ax.set_title("Theme Composites — Correlation (5y)")
                fig.colorbar(im, ax=ax, fraction=0.046, pad=0.04)
                path_corr = _save_plot(fig, outdir, "dashboard_correlation.png")
                # Revert to heatmap at the top-left
                insert_image(ws, path_corr, "A2", max_width=650, max_height=360)
            except Exception:
                pass

    # KPIs and highlights
    ws: Worksheet = wb["Dashboard"]
    try:
        reg_csv = Path("Data/processed/macro_data_with_regimes.csv")
        if reg_csv.exists():
            reg_df = pd.read_csv(reg_csv, index_col=0, parse_dates=[0])
            if "Regime_Ensemble" in reg_df.columns and len(reg_df) > 0:
                current_reg = str(reg_df["Regime_Ensemble"].iloc[-1])
                rev = reg_df["Regime_Ensemble"][::-1]
                months = int((rev == current_reg).astype(int).cumsum().iloc[0])
                ws["J2"].value = "Current Regime:"
                ws["K2"].value = current_reg
                ws["J3"].value = "Months in Regime:"
                ws["K3"].value = months
                try:
                    ws["J2"].font = Font(bold=True)
                    ws["J3"].font = Font(bold=True)
                except Exception:
                    pass
                feat_csv = Path("Data/processed/macro_data_featured.csv")
                if feat_csv.exists():
                    m = pd.read_csv(feat_csv, index_col=0, parse_dates=[0])
                    if len(m) > 24:
                        horizon = m.index.max() - pd.DateOffset(years=10)
                        base = m[m.index >= horizon] if isinstance(m.index, pd.DatetimeIndex) else m
                        z = base.apply(lambda s: (s - s.mean()) / (s.std(ddof=0) or 1.0))
                        top = z.iloc[-1].abs().sort_values(ascending=False).head(8)
                        ws["J5"].value = "Top Regime Features (|z|)"
                        try:
                            ws["J5"].font = Font(bold=True)
                        except Exception:
                            pass
                        r0 = 6
                        for i, (name, val) in enumerate(top.items()):
                            ws.cell(row=r0 + i, column=10).value = str(name)
                            ws.cell(row=r0 + i, column=11).value = float(val)

                # Model/optimizer summary box
                try:
                    ws["A6"].value = "Models in Report"
                    ws["A6"].font = Font(bold=True)
                    cols = set(reg_df.columns)
                    models = []
                    for key in ("Rule", "KMeans", "GMM", "HMM", "Supervised"):
                        if any(c for c in cols if str(c).lower().startswith(key.lower())):
                            models.append(key)
                    # Ensemble probabilities
                    prob_cols = [c for c in cols if isinstance(c, str) and c.startswith("Ensemble_Prob_")]
                    if prob_cols:
                        models.append(f"Ensemble({len(prob_cols)} probs)")
                    ws["A7"].value = ", ".join(models) if models else "n/a"

                    # Portfolio optimizer/config
                    ws["A9"].value = "Portfolio Method"
                    ws["A9"].font = Font(bold=True)
                    ws["B9"].value = "Sharpe (dynamic regime)"
                    ws["A10"].value = "Rebalance"
                    ws["B10"].value = str(REBAL_FREQ)
                    ws["A11"].value = "Lookback (yrs)"
                    ws["B11"].value = int(REGIME_WINDOW_YEARS)
                    ws["A12"].value = "Tx cost"
                    ws["B12"].value = float(TRANSACTION_COST)
                except Exception:
                    pass
    except Exception:
        pass

    # Optimal allocation and YTD chart
    try:
        reg_csv = Path("Data/processed/macro_data_with_regimes.csv")
        asset_csv = Path("Data/processed/asset_returns.csv")
        if PortfolioConstructor and reg_csv.exists() and asset_csv.exists():
            reg_df = pd.read_csv(reg_csv, index_col=0, parse_dates=[0])
            assets = pd.read_csv(asset_csv, index_col=0, parse_dates=[0])
            if "Regime_Ensemble" in reg_df.columns and not assets.empty:
                current_reg = str(reg_df["Regime_Ensemble"].iloc[-1])
                df_merged = assets.join(reg_df[["Regime_Ensemble"]], how="inner")
                df_reg = df_merged[df_merged["Regime_Ensemble"].astype(str) == current_reg].drop(columns=["Regime_Ensemble"]).dropna(how="all")
                if len(df_reg) >= 12 and df_reg.shape[1] >= 3:
                    mean_ret = pd.to_numeric(df_reg.mean(), errors="coerce")
                    cov = df_reg.cov()
                    regime_stats = {current_reg: {"mean_returns": mean_ret, "covariance": cov, "periods_per_year": 12}}
                    pc = PortfolioConstructor()
                    res = pc.optimize_portfolio(regime_stats, current_reg, method=OptimizationMethod.SHARPE)
                    weights = res.weights.sort_values(ascending=False)
                    ws["J15"].value = "Optimal Weights (current regime)"
                    try:
                        ws["J15"].font = Font(bold=True)
                    except Exception:
                        pass
                    for i, (name, val) in enumerate(weights.items(), start=16):
                        ws.cell(row=i, column=10).value = str(name)
                        ws.cell(row=i, column=11).value = float(val)
                    # Add bar chart for allocation
                    try:
                        data = Reference(ws, min_col=11, min_row=16, max_row=16 + len(weights) - 1)
                        cats = Reference(ws, min_col=10, min_row=16, max_row=16 + len(weights) - 1)
                        bch = BarChart()
                        bch.title = "Allocation Weights"
                        bch.add_data(data, titles_from_data=False)
                        bch.set_categories(cats)
                        ws.add_chart(bch, "L15")
                    except Exception:
                        pass
                    this_year = df_merged.index.max().year
                    ar = assets[assets.index.year == this_year].dropna(how="all")
                    if not ar.empty:
                        w = weights.reindex(ar.columns).fillna(0.0)
                        rp = (ar.fillna(0.0).dot(w)).add(1).cumprod() - 1
                        bench = _build_6040_benchmark(ar)
                        if bench is None:
                            bench = ar.fillna(0.0).mean(axis=1)
                        ew = bench.add(1).cumprod() - 1
                        fig, ax = plt.subplots(figsize=(7.4, 3.8))
                        rp.plot(ax=ax, label="Regime Portfolio", color="#1f77b4")
                        ew.plot(ax=ax, label="60/40 Benchmark", color="#888888")
                        ax.set_title("YTD Portfolio PnL")
                        ax.set_ylabel("Cumulative Return")
                        ax.grid(True, alpha=0.25)
                        ax.legend(loc="best")
                        p = _save_plot(fig, outdir, "dashboard_mmi.png")
                        insert_image(ws, p, "A22")
    except Exception:
        pass


def build_regime_labels(wb, outdir: Path) -> None:
    def _render(df: pd.DataFrame) -> Optional[Path]:
        try:
            regime_col = "Regime_Ensemble" if "Regime_Ensemble" in df.columns else "Regime"
            if regime_col not in df.columns:
                return None
            codes = pd.Categorical(df[regime_col]).codes
            fig, (ax1, ax2) = plt.subplots(nrows=2, ncols=1, figsize=(7.4, 4.6), sharex=True, gridspec_kw={"height_ratios": [2.0, 1.0]})
            ax1.plot(df.index, codes, drawstyle="steps-post", linewidth=1.6, color="#cc3366")
            ax1.set_title("Global Macro Regime (history)")
            ax1.set_yticks([])
            try:
                ax1.set_ylim(-0.5, float(np.nanmax(codes)) + 0.5)
            except Exception:
                pass
            ax1.grid(True, axis="x", alpha=0.25)
            prob_cols = [c for c in df.columns if isinstance(c, str) and c.startswith("Ensemble_Prob_")]
            if prob_cols:
                probs = pd.DataFrame({c: pd.to_numeric(df[c], errors="coerce") for c in prob_cols}).fillna(0.0)
                rs = probs.sum(axis=1).replace(0.0, np.nan)
                probs = probs.div(rs, axis=0).fillna(0.0).tail(180)
                probs.plot.area(ax=ax2, stacked=True, linewidth=0)
                ax2.set_ylim(0, 1)
                ax2.set_yticks([0, 0.5, 1.0])
                ax2.set_title("Ensemble Probabilities (last 180 months)")
                ax2.grid(True, axis="y", alpha=0.2)

                # Low-confidence shading (confidence < 0.6)
                try:
                    if "Ensemble_Confidence" in df.columns:
                        conf = pd.to_numeric(df["Ensemble_Confidence"], errors="coerce").fillna(0.0)
                    else:
                        conf = probs.max(axis=1)
                    conf = conf.reindex(probs.index)
                    low = conf < 0.6
                    # find contiguous low segments
                    if low.any():
                        # convert to segments of start/end indices
                        in_seg = False
                        start = None
                        dates = probs.index
                        for i, flag in enumerate(low.values):
                            if flag and not in_seg:
                                in_seg = True
                                start = dates[i]
                            if in_seg and (not flag or i == len(low.values) - 1):
                                end = dates[i] if not flag else dates[i]
                                # draw shaded region on both axes
                                ax1.axvspan(start, end, color="#000000", alpha=0.08)
                                ax2.axvspan(start, end, color="#000000", alpha=0.08)
                                in_seg = False
                except Exception:
                    pass
            else:
                ax2.text(0.01, 0.5, "No ensemble probabilities available", transform=ax2.transAxes)
                ax2.set_axis_off()
            return _save_plot(fig, outdir, "regime_history.png")
        except Exception:
            return None

    # Prefer the authoritative processed CSV for regime series to avoid
    # accidental contamination from any existing sheet contents. Fallback to
    # reading the sheet only if the CSV is missing.
    df_source: Optional[pd.DataFrame] = None
    csv_path = Path("Data/processed/macro_data_with_regimes.csv")
    if csv_path.exists():
        try:
            df_source = pd.read_csv(csv_path, index_col=0, parse_dates=[0])
        except Exception:
            df_source = None

    if df_source is None and "Regime Labels" in wb.sheetnames:
        # Fallback: attempt to reconstruct a DataFrame from current sheet
        try:
            ws_tmp: Worksheet = wb["Regime Labels"]
            data = [list(row) for row in ws_tmp.iter_rows(values_only=True)]
            if data:
                header = data[0]
                df_candidate = pd.DataFrame(data[1:], columns=header)
                try:
                    df_candidate.iloc[:, 0] = pd.to_datetime(df_candidate.iloc[:, 0])
                    df_candidate = df_candidate.set_index(df_candidate.columns[0])
                except Exception:
                    pass
                df_source = df_candidate
        except Exception:
            df_source = None

    if df_source is None or df_source.empty:
        return

    path = _render(df_source)
    if path is None:
        return

    # Replace the sheet entirely to ensure no tables/listobjects remain
    try:
        if "Regime Labels" in wb.sheetnames:
            ws_old = wb["Regime Labels"]
            wb.remove(ws_old)
        wb.create_sheet("Regime Labels")
    except Exception:
        # Fallback to clearing if remove fails
        if "Regime Labels" not in wb.sheetnames:
            wb.create_sheet("Regime Labels")
        ws_tmp2: Worksheet = wb["Regime Labels"]
        try:
            for img in list(getattr(ws_tmp2, "_images", [])):
                ws_tmp2._images.remove(img)
            max_rows = ws_tmp2.max_row or 1
            if max_rows > 0:
                ws_tmp2.delete_rows(1, max_rows)
        except Exception:
            pass
    ws: Worksheet = wb["Regime Labels"]

    insert_image(ws, path, "A2", max_width=650, max_height=360)


def build_inplace(workbook_path: str, out: Optional[str] = None) -> None:
    wb = load_workbook_safe(workbook_path)
    outdir = Path("Output/excel_charts")
    for theme in THEME_SHEETS:
        try:
            build_theme_sheet(wb, theme, outdir)
        except Exception:
            continue
    # New high-signal charts via helper builders (gracefully skipped on errors)
    try:
        featured_csv = Path("Data/processed/macro_data_featured.csv")
        macro_df: Optional[pd.DataFrame] = None
        if featured_csv.exists():
            try:
                macro_df = pd.read_csv(featured_csv, index_col=0, parse_dates=[0])
            except Exception:
                macro_df = None
        if macro_df is None:
            macro_df = load_data_dump_df(wb)
        if macro_df is not None and not macro_df.empty and build_growth_sheet:
            sheet_growth = "Growth & Labour" if "Growth & Labour" in wb.sheetnames else ("Growth" if "Growth" in wb.sheetnames else THEME_SHEETS[0])
            sheet_housing = "Housing" if "Housing" in wb.sheetnames else THEME_SHEETS[-2]
            sheet_infl = "Inflation & Liquidity" if "Inflation & Liquidity" in wb.sheetnames else ("Inflation" if "Inflation" in wb.sheetnames else THEME_SHEETS[1])
            sheet_credit = "Credit & Risk" if "Credit & Risk" in wb.sheetnames else ("Credit" if "Credit" in wb.sheetnames else THEME_SHEETS[2])
            try:
                build_growth_sheet(macro_df, wb, CHART_CFG, sheet_name=sheet_growth)
            except Exception:
                pass
            try:
                build_housing_sheet(macro_df, wb, CHART_CFG, sheet_name=sheet_housing)
            except Exception:
                pass
            try:
                build_inflation_sheet(macro_df, wb, CHART_CFG, sheet_name=sheet_infl)
            except Exception:
                pass
            try:
                build_credit_sheet(macro_df, wb, CHART_CFG, sheet_name=sheet_credit)
            except Exception:
                pass
    except Exception:
        pass
    build_dashboard(wb, outdir)
    build_regime_labels(wb, outdir)
    try:
        from glob import glob
        shots = sorted(glob("Output/portfolio_comparison_*.png"))
        if shots:
            sheetname = "Portfolios"
            if sheetname not in wb.sheetnames:
                wb.create_sheet(sheetname)
            ws: Worksheet = wb[sheetname]
            try:
                for img in list(ws._images):
                    ws._images.remove(img)
            except Exception:
                pass
            # Arrange charts in a 2-column grid with consistent scaling
            grid_cols = ["A", "J"]
            base_row = 2
            row_step = 28
            for i, shot in enumerate(shots):
                col = grid_cols[i % len(grid_cols)]
                row = base_row + (i // len(grid_cols)) * row_step
                anchor = f"{col}{row}"
                try:
                    insert_image(ws, Path(shot), anchor, max_width=440, max_height=260)
                except Exception:
                    continue
    except Exception:
        pass
    out_path = out or "Output/Macro_Reg_Report_with_category_dashboards.xlsm"
    wb.save(out_path)
    print(f"Workbook saved: {out_path}")


def main():
    import argparse
    import os
    import sys
    import subprocess

    parser = argparse.ArgumentParser(description="Build charts into Macro Reg Template workbook")
    parser.add_argument("--workbook", default="tests/Macro Reg Template.xlsm", help="Path to workbook (.xlsx or .xlsm)")
    parser.add_argument("--out", default=None, help="Optional output path (overwrite in place if omitted)")
    parser.add_argument("--open", action="store_true", help="Open the workbook after saving")
    args = parser.parse_args()
    build_inplace(args.workbook, out=args.out)

    if args.open:
        # Determine final path
        out_path = args.out or "Output/Macro_Reg_Report_with_category_dashboards.xlsm"
        try:
            if sys.platform.startswith("win"):
                os.startfile(out_path)  # type: ignore[attr-defined]
            elif sys.platform == "darwin":
                subprocess.Popen(["open", out_path])
            else:
                subprocess.Popen(["xdg-open", out_path])
        except Exception:
            print(f"Open the workbook manually: {out_path}")


if __name__ == "__main__":
    main()


