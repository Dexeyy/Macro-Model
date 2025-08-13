from __future__ import annotations

from pathlib import Path
from typing import Dict

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.chart import BarChart, Reference, LineChart


def _ensure_sheet(wb: Workbook, name: str) -> Worksheet:
    if name in wb.sheetnames:
        ws = wb[name]
        # clear
        try:
            for img in list(getattr(ws, "_images", [])):
                ws._images.remove(img)
        except Exception:
            pass
        ws.delete_rows(1, ws.max_row or 1)
        return ws
    return wb.create_sheet(name)


def build_dashboard(
    wb: Workbook,
    regime_info: Dict,
    top_features: pd.DataFrame,
    allocation: pd.Series,
    factor_heatmap: pd.DataFrame,
    pnl_df: pd.DataFrame,
) -> None:
    ws = _ensure_sheet(wb, "Dashboard")

    # Hero panel
    ws["A1"].value = "Macro Regime Dashboard"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A3"].value = "Current Regime:"
    ws["B3"].value = str(regime_info.get("label", ""))
    ws["A4"].value = "Months in Regime:"
    ws["B4"].value = int(regime_info.get("months", 0))

    # Probabilities bar (horizontal stacked using cells)
    probs = regime_info.get("proba", {}) or {}
    if probs:
        start_col = 4
        ws["A6"].value = "Regime Probabilities"
        total = sum(float(v) for v in probs.values()) or 1.0
        for i, (name, val) in enumerate(probs.items()):
            ws.cell(row=6, column=start_col + i).value = float(val) / total
        # Simple bar chart
        ch = BarChart()
        ch.type = "bar"
        ch.title = "Regime Probabilities"
        data = Reference(ws, min_col=start_col, min_row=6, max_col=start_col + len(probs) - 1, max_row=6)
        ch.add_data(data, titles_from_data=False)
        ws.add_chart(ch, "A7")

    # Top features table
    if isinstance(top_features, pd.DataFrame) and not top_features.empty:
        ws["A16"].value = "Top Features (z-score)"
        ws["A16"].font = Font(bold=True)
        tbl = top_features.head(5).copy()
        ws["A17"].value = "Feature"
        ws["B17"].value = "z"
        for i, (name, row) in enumerate(tbl.iterrows(), start=18):
            ws.cell(row=i, column=1).value = str(name)
            ws.cell(row=i, column=2).value = float(row.iloc[0]) if len(row) else float(row)

    # Allocation table + bar chart
    if isinstance(allocation, pd.Series) and not allocation.empty:
        ws["D16"].value = "Optimal Weights (current regime)"
        ws["D16"].font = Font(bold=True)
        ws["D17"].value = "Asset"
        ws["E17"].value = "Weight"
        for i, (name, w) in enumerate(allocation.items(), start=18):
            ws.cell(row=i, column=4).value = str(name)
            ws.cell(row=i, column=5).value = float(w)
        ch = BarChart()
        ch.title = "Allocation"
        data = Reference(ws, min_col=5, min_row=18, max_row=18 + len(allocation) - 1)
        cats = Reference(ws, min_col=4, min_row=18, max_row=18 + len(allocation) - 1)
        ch.add_data(data, titles_from_data=False)
        ch.set_categories(cats)
        ws.add_chart(ch, "G16")

    # Factor heatmap (as values table to be formatted by Excel)
    if isinstance(factor_heatmap, pd.DataFrame) and not factor_heatmap.empty:
        ws["A24"].value = "Factor Heatmap (last 24 months)"
        ws["A24"].font = Font(bold=True)
        mat = factor_heatmap.tail(24)
        ws["A25"].value = "Date"
        for j, c in enumerate(mat.columns, start=2):
            ws.cell(row=25, column=j).value = str(c)
        for i, (dt, row) in enumerate(mat.iterrows(), start=26):
            ws.cell(row=i, column=1).value = dt
            for j, v in enumerate(row.values, start=2):
                ws.cell(row=i, column=j).value = float(v) if pd.notna(v) else None

    # YTD PnL line chart
    if isinstance(pnl_df, pd.DataFrame) and not pnl_df.empty:
        ws["A50"].value = "YTD Regime vs Benchmark"
        ws["A50"].font = Font(bold=True)
        base_row = 51
        ws.cell(row=base_row, column=1).value = "Date"
        for j, c in enumerate(pnl_df.columns, start=2):
            ws.cell(row=base_row, column=j).value = str(c)
        for i, (dt, row) in enumerate(pnl_df.iterrows(), start=base_row + 1):
            ws.cell(row=i, column=1).value = dt
            for j, v in enumerate(row.values, start=2):
                ws.cell(row=i, column=j).value = float(v)
        ch = LineChart()
        ch.title = "YTD PnL"
        data = Reference(ws, min_col=2, min_row=base_row, max_col=1 + len(pnl_df.columns), max_row=base_row + len(pnl_df))
        ch.add_data(data, titles_from_data=True)
        ws.add_chart(ch, "G48")

        # Metrics table (annualized return, vol, Sharpe, max DD)
        try:
            rp = pnl_df.iloc[:, 0].astype(float)
            ann = 12
            ann_ret = (1 + rp.mean()) ** ann - 1
            ann_vol = rp.std() * (ann ** 0.5)
            sharpe = ann_ret / ann_vol if ann_vol > 0 else 0.0
            cum = (1 + rp).cumprod()
            dd = (cum / cum.cummax()) - 1
            mdd = float(dd.min()) if len(dd) else 0.0
            ws["J48"].value = "Regime Portfolio Metrics"
            ws["J48"].font = Font(bold=True)
            ws["J49"].value = "Ann. Return"
            ws["K49"].value = float(ann_ret)
            ws["J50"].value = "Ann. Vol"
            ws["K50"].value = float(ann_vol)
            ws["J51"].value = "Sharpe"
            ws["K51"].value = float(sharpe)
            ws["J52"].value = "Max DD"
            ws["K52"].value = float(mdd)
        except Exception:
            pass


