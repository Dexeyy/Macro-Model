import json
from pathlib import Path

import pandas as pd

from tools.curate_charts import curate_plan
from src.reporting.dashboard_writer import build_dashboard
from openpyxl import Workbook
from src.models.portfolio import compute_dynamic_regime_portfolio, OptimizationMethod


def test_curate_drops_duplicates(tmp_path: Path):
    # Fake charts.csv with duplicate series on same sheet
    charts_csv = tmp_path / "charts.csv"
    rows = [
        ["Growth & Labour", "Chart A", "INDPRO", "=Sheet!A1:A10"],
        ["Growth & Labour", "Chart B", "INDPRO", "=Sheet!A1:A10"],
        ["Growth & Labour", "Chart C", "PAYEMS", "=Sheet!B1:B10"],
    ]
    charts_csv.write_text("sheet,title,series_titles,series_formulas\n" + "\n".join([",".join(r) for r in rows]))
    idx = pd.date_range("2020-01-01", periods=60, freq="M")
    monthly = pd.DataFrame({"INDPRO": range(60), "PAYEMS": range(60)}, index=idx)
    monthly_csv = tmp_path / "monthly.csv"
    monthly.to_csv(monthly_csv)
    plan = curate_plan(charts_csv, monthly_csv)
    assert "Growth & Labour" in plan["keep"]
    assert len(plan["keep"]["Growth & Labour"]) >= 1
    assert any("Chart B" in d for d in plan["drop"]["Growth & Labour"]) or any(
        "Chart A" in d for d in plan["drop"]["Growth & Labour"]
    )


def test_dashboard_writer_inserts_sections(tmp_path: Path):
    wb = Workbook()
    regime = {"label": "Expansion", "months": 8, "proba": {"P0": 0.6, "P1": 0.3, "P2": 0.1}}
    top = pd.DataFrame({"z": [2.1, -1.5, 1.2]}, index=["INDPRO", "UNRATE", "PAYEMS"])
    alloc = pd.Series({"SPX": 0.6, "US10Y_NOTE_FUT": 0.4})
    heat = pd.DataFrame({"F_Growth": [0.1, 0.2]}, index=pd.date_range("2023-01-01", periods=2, freq="M"))
    pnl = pd.DataFrame({"Regime Portfolio": [0.0, 0.02], "Equal-Weight": [0.0, 0.01]}, index=pd.date_range("2024-01-01", periods=2, freq="M"))
    build_dashboard(wb, regime, top, alloc, heat, pnl)
    assert "Dashboard" in wb.sheetnames
    ws = wb["Dashboard"]
    assert ws["A1"].value == "Macro Regime Dashboard"


def test_dynamic_regime_portfolio_rebalances_on_change():
    idx = pd.date_range("2023-01-31", periods=12, freq="M")
    # Two assets, switch regime mid-year
    ret = pd.DataFrame({"A": [0.01]*6 + [0.00]*6, "B": [0.00]*6 + [0.01]*6}, index=idx)
    reg = pd.Series(["R1"]*6 + ["R2"]*6, index=idx)
    series, w_hist = compute_dynamic_regime_portfolio(
        ret,
        reg,
        regime_window_years=1,
        method=OptimizationMethod.SHARPE,
        rebal_freq="M",
        transaction_cost=0.0,
    )
    # Expect at least two rebalances
    assert len(w_hist) >= 2
    # Weights should shift from A to B across regimes
    first_w = next(iter(w_hist.values()))
    last_w = list(w_hist.values())[-1]
    assert first_w.idxmax() != last_w.idxmax()


